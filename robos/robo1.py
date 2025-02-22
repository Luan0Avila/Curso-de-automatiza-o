from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

print("iniciando robo....\n")
arq = open("resultado.txt", "w")

dominios = []
#Lendo do excel
workbook = load_workbook('./robos/dominio.xlsx')
sheet = workbook.active

for linha in range(1,10):
    dominios.append(sheet.cell(row=linha, column=1).value)


driver = webdriver.Chrome()
driver.get("https://registro.br/")


for dominio in dominios:
    pesquisa = driver.find_element(By.NAME,"is-avail")
    pesquisa.clear() #limpa a barra de pesquisa
    pesquisa.send_keys(dominio)
    pesquisa.send_keys(Keys.RETURN)
    time.sleep(3)
    resultados = driver.find_elements(By.TAG_NAME, "strong")
    texto = "Dominio %s %s" % (dominio, resultados[2].text)
    arq.write(texto)


arq.close()
driver.close()